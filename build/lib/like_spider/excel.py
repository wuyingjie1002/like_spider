import time, os
from openpyxl import load_workbook
from openpyxl import Workbook
from .config import *

class Excel():
    """This is a class that saves data to an excel file."""

    def loadFile(self, fileName):
        """load excel file"""
        self.wb = load_workbook(fileName)
        self.sheets = self.wb.get_sheet_names()

    def loadSheet(self, sheet):
        """load a sheet"""
        self.table = self.wb[sheet]
        self.rows = self.table.max_row
        self.cols = self.table.max_column

    def getValue(self, row, col):
        """get a value"""
        return self.table.cell(row, col).value

    def saveFile(self, data, fileName):
        """save data to an excel file."""
        if fileName == "":
            print('file error')
            return False
        totalRow = len(data)
        if totalRow > 0:
            wb = Workbook()
            ws = wb.active
            for row in range(1, (totalRow + 1)):
                totalCol = len(data[(row - 1)])
                if totalCol > 0:
                    for col in range(1, (totalCol + 1)):
                        cell = ws.cell(row = row, column = col)
                        cell.value = data[(row - 1)][(col - 1)]
                else:
                    print('col data error')
                    break
            if totalCol > 0:
                wb.save(fileName)
        else:
            print('row data error')

    def appendFile(self, data, fileName, sheet = ''):
        """append data to an excel file."""
        if fileName == "":
            print('file error')
            return False
        if os.path.exists(fileName):
            self.loadFile(fileName)
            if sheet == '':
                sheet = self.sheets[0]
            self.loadSheet(sheet)
            if self.rows > 0 and self.cols > 0:
                fileData = []
                for row in range(1, self.rows + 1):
                    rowData = []
                    for col in range(1, self.cols + 1):
                        rowData.append(self.getValue(row, col))
                    fileData.append(rowData)
                fileData.extend(data)
                data = fileData
        self.saveFile(data, fileName)